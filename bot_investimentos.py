import csv
import io
import logging
import os
import re
import threading
import unicodedata
import zipfile
from collections import defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from difflib import get_close_matches
from http.server import BaseHTTPRequestHandler, HTTPServer

import psycopg2
import psycopg2.extras
from telegram import InputFile, ReplyKeyboardMarkup, ReplyKeyboardRemove, Update
from telegram.ext import (
    Application,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)


logging.basicConfig(level=logging.WARNING)

TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN") or os.environ.get("BOT_TOKEN")
DATABASE_URL = os.environ.get("DATABASE_URL")

MONEY = Decimal("0.01")
QTY = Decimal("0.00000001")

CANCELAR_BTN = "❌ Cancelar"
VOLTAR_BTN = "⬅️ Voltar"
CONFIRMAR_NOVO_PREFIX = "✅ Confirmar novo:"
SALVAR_COMPRA_BTN = "✅ Salvar compra"
SALVAR_VENDA_BTN = "✅ Salvar venda"
SALVAR_PROVENTO_BTN = "✅ Salvar provento"
SIM_DELETAR_BTN = "SIM, deletar"
NAO_DELETAR_BTN = "Não deletar"

CORRETORAS = [
    "XP",
    "Rico",
    "Clear",
    "Nubank",
    "Inter",
    "BTG Pactual",
    "Avenue",
    "Toro",
    "Genial",
    "Modalmais",
    "Itaú",
    "Bradesco",
    "Santander",
    "Banco do Brasil",
    "Caixa",
    "Warren",
    "Órama",
    "Guide",
    "Binance",
    "Mercado Bitcoin",
    "Outra",
]

TIPOS_INVESTIMENTO = [
    "Ação",
    "FII",
    "Renda Fixa",
    "Tesouro Direto",
    "Criptomoeda",
    "ETF",
    "Fundo",
    "Outro",
]

TIPOS_PROVENTO = ["Dividendo", "JCP", "Rendimento"]

MAPA_CORRETORAS = {
    "xp": "XP",
    "xp investimentos": "XP",
    "xpi": "XP",
    "rico": "Rico",
    "rico investimentos": "Rico",
    "clear": "Clear",
    "clear corretora": "Clear",
    "nubank": "Nubank",
    "nu invest": "Nubank",
    "nuinvest": "Nubank",
    "easynvest": "Nubank",
    "inter": "Inter",
    "inter invest": "Inter",
    "banco inter": "Inter",
    "btg": "BTG Pactual",
    "btg pactual": "BTG Pactual",
    "btg investimentos": "BTG Pactual",
    "avenue": "Avenue",
    "avenue securities": "Avenue",
    "toro": "Toro",
    "toro investimentos": "Toro",
    "genial": "Genial",
    "genial investimentos": "Genial",
    "modal": "Modalmais",
    "modalmais": "Modalmais",
    "itau": "Itaú",
    "itaú": "Itaú",
    "itau corretora": "Itaú",
    "bradesco": "Bradesco",
    "bradesco corretora": "Bradesco",
    "santander": "Santander",
    "bb": "Banco do Brasil",
    "banco do brasil": "Banco do Brasil",
    "caixa": "Caixa",
    "warren": "Warren",
    "orama": "Órama",
    "órama": "Órama",
    "guide": "Guide",
    "binance": "Binance",
    "mercado bitcoin": "Mercado Bitcoin",
    "mb": "Mercado Bitcoin",
}

MAPA_ATIVOS = {
    "petr4": "PETR4",
    "petrobras pn": "PETR4",
    "petroleo brasileiro pn": "PETR4",
    "petroleo brasileiro preferencial": "PETR4",
    "petr3": "PETR3",
    "petrobras on": "PETR3",
    "petroleo brasileiro on": "PETR3",
    "vale3": "VALE3",
    "vale on": "VALE3",
    "itub4": "ITUB4",
    "itau unibanco pn": "ITUB4",
    "itaú unibanco pn": "ITUB4",
    "itub3": "ITUB3",
    "b3sa3": "B3SA3",
    "b3": "B3SA3",
    "bbas3": "BBAS3",
    "banco do brasil on": "BBAS3",
    "bbdc4": "BBDC4",
    "bradesco pn": "BBDC4",
    "abev3": "ABEV3",
    "ambev": "ABEV3",
    "mglu3": "MGLU3",
    "magalu": "MGLU3",
    "weg3": "WEGE3",
    "wege3": "WEGE3",
    "weg": "WEGE3",
    "mxrf11": "MXRF11",
    "mxrf": "MXRF11",
    "maxi renda": "MXRF11",
    "hglg11": "HGLG11",
    "knri11": "KNRI11",
    "xpml11": "XPML11",
    "bova11": "BOVA11",
    "ivvb11": "IVVB11",
    "smal11": "SMAL11",
    "hash11": "HASH11",
    "btc": "BTC",
    "bitcoin": "BTC",
    "eth": "ETH",
    "ethereum": "ETH",
    "sol": "SOL",
    "solana": "SOL",
    "usdt": "USDT",
    "tether": "USDT",
}

SCHEMA_SQL = """
BEGIN;

CREATE TABLE IF NOT EXISTS compras (
    id SERIAL PRIMARY KEY,
    data DATE NOT NULL,
    ativo VARCHAR(80) NOT NULL,
    tipo_investimento VARCHAR(30) NOT NULL,
    corretora VARCHAR(60) NOT NULL,
    quantidade NUMERIC(20,8) NOT NULL CHECK (quantidade > 0),
    preco_unitario NUMERIC(18,6) NOT NULL CHECK (preco_unitario >= 0),
    valor_total NUMERIC(18,2) NOT NULL CHECK (valor_total >= 0),
    observacao TEXT,
    criado_em TIMESTAMP NOT NULL DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS vendas (
    id SERIAL PRIMARY KEY,
    data DATE NOT NULL,
    ativo VARCHAR(80) NOT NULL,
    tipo_investimento VARCHAR(30),
    corretora VARCHAR(60) NOT NULL,
    quantidade NUMERIC(20,8) NOT NULL CHECK (quantidade > 0),
    preco_unitario NUMERIC(18,6) NOT NULL CHECK (preco_unitario >= 0),
    valor_total NUMERIC(18,2) NOT NULL CHECK (valor_total >= 0),
    custo_medio_unitario NUMERIC(18,6) NOT NULL DEFAULT 0,
    custo_total NUMERIC(18,2) NOT NULL DEFAULT 0,
    lucro_prejuizo NUMERIC(18,2) NOT NULL DEFAULT 0,
    criado_em TIMESTAMP NOT NULL DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS proventos (
    id SERIAL PRIMARY KEY,
    data DATE NOT NULL,
    ativo VARCHAR(80) NOT NULL,
    corretora VARCHAR(60) NOT NULL,
    tipo_provento VARCHAR(20) NOT NULL,
    valor_recebido NUMERIC(18,2) NOT NULL CHECK (valor_recebido >= 0),
    criado_em TIMESTAMP NOT NULL DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS configuracoes (
    chave VARCHAR(50) PRIMARY KEY,
    valor TEXT NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_compras_ativo_data ON compras (ativo, data, id);
CREATE INDEX IF NOT EXISTS idx_vendas_ativo_data ON vendas (ativo, data, id);
CREATE INDEX IF NOT EXISTS idx_proventos_ativo_data ON proventos (ativo, data, id);
CREATE INDEX IF NOT EXISTS idx_compras_corretora ON compras (corretora);
CREATE INDEX IF NOT EXISTS idx_vendas_corretora ON vendas (corretora);
CREATE INDEX IF NOT EXISTS idx_proventos_corretora ON proventos (corretora);

COMMIT;
""".strip()


(
    COMPRA_DATA,
    COMPRA_ATIVO,
    COMPRA_ATIVO_CONFIRM,
    COMPRA_TIPO,
    COMPRA_CORRETORA,
    COMPRA_CORRETORA_CUSTOM,
    COMPRA_QUANTIDADE,
    COMPRA_PRECO,
    COMPRA_CONFIRMAR,
    VENDA_DATA,
    VENDA_ATIVO,
    VENDA_ATIVO_CONFIRM,
    VENDA_QUANTIDADE,
    VENDA_PRECO,
    VENDA_CORRETORA,
    VENDA_CORRETORA_CUSTOM,
    VENDA_CONFIRMAR,
    PROVENTO_DATA,
    PROVENTO_ATIVO,
    PROVENTO_ATIVO_CONFIRM,
    PROVENTO_CORRETORA,
    PROVENTO_CORRETORA_CUSTOM,
    PROVENTO_TIPO,
    PROVENTO_VALOR,
    PROVENTO_CONFIRMAR,
    EDITAR_TIPO,
    EDITAR_PAGINA,
    EDITAR_CAMPO,
    EDITAR_VALOR,
    EDITAR_SELECT,
    RELATORIOS_MENU,
    GERAR_MENU,
) = range(32)


def conectar():
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL não definida nas variáveis de ambiente.")
    return psycopg2.connect(DATABASE_URL)


def inicializar_db():
    with conectar() as conn:
        with conn.cursor() as cur:
            cur.execute(SCHEMA_SQL)
        conn.commit()


def texto_chave(valor):
    valor = (valor or "").strip().lower()
    valor = unicodedata.normalize("NFKD", valor)
    valor = "".join(ch for ch in valor if not unicodedata.combining(ch))
    valor = re.sub(r"\s+", " ", valor)
    return valor


def limpar_ticker(valor):
    valor = (valor or "").strip().upper()
    valor = re.sub(r"\.SA$", "", valor)
    valor = valor.replace("-", "").replace("_", "")
    return re.sub(r"[^A-Z0-9]", "", valor)


def normalizar_ativo(raw):
    raw = (raw or "").strip()
    if not raw:
        return ""
    chave = texto_chave(raw)
    if chave in MAPA_ATIVOS:
        return MAPA_ATIVOS[chave]
    ticker = limpar_ticker(raw)
    if re.fullmatch(r"[A-Z]{2,6}\d{1,2}[A-Z]?", ticker) or re.fullmatch(r"[A-Z]{2,10}", ticker):
        return ticker
    sem_pontuacao = re.sub(r"[^A-Za-zÀ-ÿ0-9 ]", " ", raw)
    return re.sub(r"\s+", " ", sem_pontuacao).strip().upper()


def normalizar_corretora(raw):
    raw = (raw or "").strip()
    if not raw:
        return ""
    chave = texto_chave(raw)
    if chave in MAPA_CORRETORAS:
        return MAPA_CORRETORAS[chave]
    for corretora in CORRETORAS:
        if texto_chave(corretora) == chave:
            return corretora
    return raw.title()


def moeda(valor):
    return Decimal(valor or 0).quantize(MONEY, rounding=ROUND_HALF_UP)


def qtd(valor):
    return Decimal(valor or 0).quantize(QTY, rounding=ROUND_HALF_UP)


def parse_decimal(raw):
    raw = (raw or "").strip()
    raw = raw.replace("R$", "").replace(" ", "")
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    else:
        raw = raw.replace(",", ".")
    try:
        return Decimal(raw)
    except (InvalidOperation, ValueError):
        return None


def fmt_decimal(valor, decimais=2):
    valor = Decimal(valor or 0)
    texto = f"{valor:.{decimais}f}"
    if decimais > 0:
        texto = texto.rstrip("0").rstrip(".")
    return texto.replace(".", ",")


def fmt_moeda(valor):
    valor = moeda(valor)
    sinal = "-" if valor < 0 else ""
    valor_abs = abs(valor)
    inteiro, centavos = f"{valor_abs:.2f}".split(".")
    grupos = []
    while inteiro:
        grupos.insert(0, inteiro[-3:])
        inteiro = inteiro[:-3]
    return f"{sinal}R$ {'.'.join(grupos)},{centavos}"


def data_obj(valor):
    if hasattr(valor, "strftime"):
        return valor
    return datetime.strptime(str(valor)[:10], "%Y-%m-%d").date()


def data_br(valor):
    return data_obj(valor).strftime("%d/%m/%Y")


def normalizar_data(raw):
    agora = datetime.now()
    raw = (raw or "").strip()
    if raw == "0":
        return agora.strftime("%Y-%m-%d")
    try:
        dia = int(raw)
        if 1 <= dia <= 31:
            return datetime(agora.year, agora.month, dia).strftime("%Y-%m-%d")
    except ValueError:
        pass
    for fmt in ("%d/%m", "%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            data = datetime.strptime(raw, fmt)
            if fmt == "%d/%m":
                data = data.replace(year=agora.year)
            return data.strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def teclado_menu():
    return ReplyKeyboardMarkup(
        [
            ["🟢 Registrar Compra", "🔴 Registrar Venda"],
            ["💰 Registrar Provento", "📈 Relatórios"],
            ["✏️ Editar Registro", "📊 Gerar Arquivos"],
        ],
        resize_keyboard=True,
    )


def teclado_cancelar():
    return ReplyKeyboardMarkup([[CANCELAR_BTN]], resize_keyboard=True)


def teclado_confirmar(botao):
    return ReplyKeyboardMarkup([[botao], [CANCELAR_BTN]], resize_keyboard=True, one_time_keyboard=True)


def teclado_lista(lista, incluir_outra=False):
    itens = [x for x in lista if x != "Outra"]
    linhas = [[x] for x in itens]
    if incluir_outra:
        linhas.append(["Outra"])
    linhas.append([CANCELAR_BTN])
    return ReplyKeyboardMarkup(linhas, resize_keyboard=True, one_time_keyboard=True)


def teclado_gerar():
    return ReplyKeyboardMarkup(
        [["📊 Dashboard Excel"], ["📂 Exportar CSV"], ["🧱 Script SQL"], ["🔙 Voltar"]],
        resize_keyboard=True,
    )


def teclado_relatorios():
    return ReplyKeyboardMarkup(
        [
            ["📌 Posição Atual", "🏦 Por Corretora"],
            ["🏷 Por Tipo", "📉 Vendas"],
            ["💰 Proventos", "📅 Evolução"],
            ["🔙 Voltar"],
        ],
        resize_keyboard=True,
    )


def fetch_all(sql, params=()):
    with conectar() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params)
            return [dict(r) for r in cur.fetchall()]


def fetch_one(sql, params=()):
    with conectar() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params)
            row = cur.fetchone()
            return dict(row) if row else None


def listar_compras():
    return fetch_all("SELECT * FROM compras ORDER BY data ASC, id ASC")


def listar_vendas():
    return fetch_all("SELECT * FROM vendas ORDER BY data ASC, id ASC")


def listar_proventos():
    return fetch_all("SELECT * FROM proventos ORDER BY data ASC, id ASC")


def listar_ativos_existentes():
    rows = fetch_all(
        """
        SELECT ativo FROM compras
        UNION
        SELECT ativo FROM vendas
        UNION
        SELECT ativo FROM proventos
        ORDER BY ativo
        """
    )
    return [r["ativo"] for r in rows if r.get("ativo")]


def get_corretoras_ordenadas(n=12):
    try:
        rows = fetch_all(
            """
            SELECT corretora, COUNT(*) AS total
            FROM (
                SELECT corretora FROM compras
                UNION ALL SELECT corretora FROM vendas
                UNION ALL SELECT corretora FROM proventos
            ) x
            WHERE corretora IS NOT NULL AND corretora <> ''
            GROUP BY corretora
            ORDER BY total DESC
            LIMIT %s
            """,
            (n,),
        )
        top = [r["corretora"] for r in rows if r["corretora"] and r["corretora"] != "Outra"]
        for corretora in CORRETORAS:
            if corretora not in top and corretora != "Outra" and len(top) < n:
                top.append(corretora)
        return top + ["Outra"]
    except Exception:
        return CORRETORAS


def tipo_do_ativo(ativo):
    row = fetch_one(
        """
        SELECT tipo_investimento
        FROM compras
        WHERE ativo = %s
        ORDER BY data DESC, id DESC
        LIMIT 1
        """,
        (ativo,),
    )
    return row["tipo_investimento"] if row else None


def inserir_compra(dados):
    with conectar() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO compras
                    (data, ativo, tipo_investimento, corretora, quantidade, preco_unitario, valor_total, observacao)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                RETURNING id
                """,
                (
                    dados["data"],
                    dados["ativo"],
                    dados["tipo_investimento"],
                    dados["corretora"],
                    dados["quantidade"],
                    dados["preco_unitario"],
                    dados["valor_total"],
                    dados.get("observacao"),
                ),
            )
            novo_id = cur.fetchone()[0]
        conn.commit()
    recalcular_vendas([dados["ativo"]])
    return novo_id


def inserir_venda(dados):
    tipo = tipo_do_ativo(dados["ativo"])
    with conectar() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO vendas
                    (data, ativo, tipo_investimento, corretora, quantidade, preco_unitario,
                     valor_total, custo_medio_unitario, custo_total, lucro_prejuizo)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                RETURNING id
                """,
                (
                    dados["data"],
                    dados["ativo"],
                    tipo,
                    dados["corretora"],
                    dados["quantidade"],
                    dados["preco_unitario"],
                    dados["valor_total"],
                    dados["custo_medio_unitario"],
                    dados["custo_total"],
                    dados["lucro_prejuizo"],
                ),
            )
            novo_id = cur.fetchone()[0]
        conn.commit()
    recalcular_vendas([dados["ativo"]])
    return novo_id


def inserir_provento(dados):
    with conectar() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO proventos
                    (data, ativo, corretora, tipo_provento, valor_recebido)
                VALUES (%s,%s,%s,%s,%s)
                RETURNING id
                """,
                (
                    dados["data"],
                    dados["ativo"],
                    dados["corretora"],
                    dados["tipo_provento"],
                    dados["valor_recebido"],
                ),
            )
            novo_id = cur.fetchone()[0]
        conn.commit()
    return novo_id


def eventos_do_ativo(ativo):
    compras = fetch_all("SELECT *, 'compra' AS evento FROM compras WHERE ativo=%s", (ativo,))
    vendas = fetch_all("SELECT *, 'venda' AS evento FROM vendas WHERE ativo=%s", (ativo,))
    eventos = compras + vendas
    return sorted(eventos, key=lambda r: (data_obj(r["data"]), int(r["id"])))


def recalcular_vendas(ativos=None):
    if ativos is None:
        ativos = listar_ativos_existentes()
    if isinstance(ativos, str):
        ativos = [ativos]
    updates = []
    for ativo in set(a for a in ativos if a):
        quantidade_atual = Decimal("0")
        custo_total_atual = Decimal("0")
        tipo_atual = tipo_do_ativo(ativo)
        for evento in eventos_do_ativo(ativo):
            if evento["evento"] == "compra":
                quantidade_atual += Decimal(evento["quantidade"])
                custo_total_atual += Decimal(evento["valor_total"])
                tipo_atual = evento.get("tipo_investimento") or tipo_atual
                continue
            quantidade_vendida = Decimal(evento["quantidade"])
            if quantidade_vendida > quantidade_atual:
                raise ValueError(
                    f"A venda #{evento['id']} de {ativo} excede a posição disponível "
                    f"({fmt_decimal(quantidade_atual, 8)} disponível)."
                )
            custo_medio = Decimal("0") if quantidade_atual == 0 else custo_total_atual / quantidade_atual
            custo_total = moeda(custo_medio * quantidade_vendida)
            valor_total = Decimal(evento["valor_total"])
            lucro = moeda(valor_total - custo_total)
            updates.append((custo_medio, custo_total, lucro, tipo_atual, evento["id"]))
            quantidade_atual -= quantidade_vendida
            custo_total_atual -= custo_total
            if quantidade_atual == 0:
                custo_total_atual = Decimal("0")
    if updates:
        with conectar() as conn:
            with conn.cursor() as cur:
                cur.executemany(
                    """
                    UPDATE vendas
                    SET custo_medio_unitario=%s, custo_total=%s, lucro_prejuizo=%s, tipo_investimento=%s
                    WHERE id=%s
                    """,
                    updates,
                )
            conn.commit()


def calcular_posicoes():
    pos = {}
    vendas_realizadas = defaultdict(Decimal)
    proventos_por_ativo = defaultdict(Decimal)
    for provento in listar_proventos():
        proventos_por_ativo[provento["ativo"]] += Decimal(provento["valor_recebido"])
    for ativo in listar_ativos_existentes():
        quantidade_atual = Decimal("0")
        custo_total_atual = Decimal("0")
        tipo_atual = tipo_do_ativo(ativo) or ""
        for evento in eventos_do_ativo(ativo):
            if evento["evento"] == "compra":
                quantidade_atual += Decimal(evento["quantidade"])
                custo_total_atual += Decimal(evento["valor_total"])
                tipo_atual = evento.get("tipo_investimento") or tipo_atual
            else:
                quantidade_vendida = Decimal(evento["quantidade"])
                if quantidade_atual <= 0:
                    continue
                custo_medio = custo_total_atual / quantidade_atual
                custo_total = moeda(custo_medio * quantidade_vendida)
                quantidade_atual -= quantidade_vendida
                custo_total_atual -= custo_total
                vendas_realizadas[ativo] += Decimal(evento.get("lucro_prejuizo") or 0)
                if quantidade_atual == 0:
                    custo_total_atual = Decimal("0")
        proventos = proventos_por_ativo[ativo]
        custo_medio_atual = Decimal("0") if quantidade_atual == 0 else custo_total_atual / quantidade_atual
        pos[ativo] = {
            "ativo": ativo,
            "tipo_investimento": tipo_atual,
            "quantidade": qtd(quantidade_atual),
            "custo_total": moeda(custo_total_atual),
            "custo_medio": custo_medio_atual,
            "lucro_realizado": moeda(vendas_realizadas[ativo]),
            "proventos": moeda(proventos),
        }
    return pos


def posicao_ativo_ate(ativo, data_limite, venda_id_excluida=None):
    data_limite_obj = datetime.strptime(str(data_limite)[:10], "%Y-%m-%d").date()
    quantidade_atual = Decimal("0")
    custo_total_atual = Decimal("0")
    for evento in eventos_do_ativo(ativo):
        data_evento = data_obj(evento["data"])
        if data_evento > data_limite_obj:
            continue
        if evento["evento"] == "compra":
            quantidade_atual += Decimal(evento["quantidade"])
            custo_total_atual += Decimal(evento["valor_total"])
            continue
        if venda_id_excluida and int(evento["id"]) == int(venda_id_excluida):
            continue
        quantidade_vendida = Decimal(evento["quantidade"])
        if quantidade_atual <= 0:
            continue
        custo_medio = custo_total_atual / quantidade_atual
        custo_total = moeda(custo_medio * quantidade_vendida)
        quantidade_atual -= quantidade_vendida
        custo_total_atual -= custo_total
        if quantidade_atual == 0:
            custo_total_atual = Decimal("0")
    custo_medio = Decimal("0") if quantidade_atual == 0 else custo_total_atual / quantidade_atual
    return qtd(quantidade_atual), moeda(custo_total_atual), custo_medio


def calcular_venda_preview(ativo, data, quantidade, preco_unitario):
    disponivel, custo_total_atual, custo_medio = posicao_ativo_ate(ativo, data)
    if quantidade > disponivel:
        return None, disponivel
    valor_total = moeda(quantidade * preco_unitario)
    custo_total = moeda(quantidade * custo_medio)
    lucro = moeda(valor_total - custo_total)
    return {
        "valor_total": valor_total,
        "custo_medio_unitario": custo_medio,
        "custo_total": custo_total,
        "lucro_prejuizo": lucro,
    }, disponivel


def sugestoes_ativos(ativo, existentes):
    if not existentes:
        return []
    matches = get_close_matches(ativo, existentes, n=5, cutoff=0.35)
    if matches:
        return matches
    alvo = texto_chave(ativo)
    ordenados = sorted(
        existentes,
        key=lambda x: abs(len(texto_chave(x)) - len(alvo)) + (0 if alvo[:3] in texto_chave(x) else 3),
    )
    return ordenados[:5]


async def resolver_ativo(update, ctx, raw, estado_confirmacao, proximo_estado, prompt_proximo):
    ativo = normalizar_ativo(raw)
    if not ativo:
        await update.message.reply_text("Digite um ativo válido:")
        return None
    existentes = listar_ativos_existentes()
    if ativo in existentes:
        ctx.user_data["ativo"] = ativo
        await prompt_proximo(update, ctx)
        return proximo_estado
    sugestoes = sugestoes_ativos(ativo, existentes)
    ctx.user_data["ativo_candidato"] = ativo
    ctx.user_data["ativo_sugestoes"] = sugestoes
    ctx.user_data["ativo_confirm_estado"] = estado_confirmacao
    ctx.user_data["ativo_proximo_estado"] = proximo_estado
    teclado = [[f"{CONFIRMAR_NOVO_PREFIX} {ativo}"]]
    teclado += [[s] for s in sugestoes]
    teclado.append([CANCELAR_BTN])
    lista = ", ".join(sugestoes) if sugestoes else "nenhum ativo parecido encontrado"
    await update.message.reply_text(
        f"Não encontrei *{ativo}* na carteira.\n\n"
        f"É um ativo novo ou você quis dizer um destes já cadastrados: {lista}?",
        reply_markup=ReplyKeyboardMarkup(teclado, resize_keyboard=True, one_time_keyboard=True),
        parse_mode="Markdown",
    )
    return estado_confirmacao


async def confirmar_ativo_generico(update, ctx, prompt_proximo):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    candidato = ctx.user_data.get("ativo_candidato")
    sugestoes = ctx.user_data.get("ativo_sugestoes", [])
    if raw.startswith(CONFIRMAR_NOVO_PREFIX):
        ctx.user_data["ativo"] = candidato
    elif raw in sugestoes:
        ctx.user_data["ativo"] = raw
    else:
        await update.message.reply_text("Escolha uma opção do teclado:")
        return ctx.user_data.get("ativo_confirm_estado")
    await prompt_proximo(update, ctx)
    return ctx.user_data["ativo_proximo_estado"]


async def voltar_menu(update, ctx):
    ctx.user_data.clear()
    await update.message.reply_text("O que mais?", reply_markup=teclado_menu())
    return ConversationHandler.END


async def cancelar(update, ctx):
    ctx.user_data.clear()
    await update.message.reply_text("Cancelado.", reply_markup=teclado_menu())
    return ConversationHandler.END


async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    await update.message.reply_text(
        "👋 *Gestor de Investimentos*\n\nEscolha uma opção:",
        reply_markup=teclado_menu(),
        parse_mode="Markdown",
    )


async def pedir_data(update, titulo):
    hoje = datetime.now().strftime("%d/%m/%Y")
    await update.message.reply_text(
        f"{titulo}\nHoje é *{hoje}*. Envie *0* para usar hoje ou digite outra data:",
        reply_markup=teclado_cancelar(),
        parse_mode="Markdown",
    )


async def compra_inicio(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    await pedir_data(update, "📅 *Data da compra*")
    return COMPRA_DATA


async def compra_data(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    data = normalizar_data(raw)
    if not data:
        await update.message.reply_text("Data inválida. Tente novamente:")
        return COMPRA_DATA
    ctx.user_data["data"] = data
    await update.message.reply_text("🏷 *Ticker ou nome do ativo:*", reply_markup=teclado_cancelar(), parse_mode="Markdown")
    return COMPRA_ATIVO


async def prompt_compra_tipo(update, ctx):
    await update.message.reply_text(
        f"📚 *Tipo de investimento para {ctx.user_data['ativo']}:*",
        reply_markup=teclado_lista(TIPOS_INVESTIMENTO),
        parse_mode="Markdown",
    )


async def compra_ativo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    return await resolver_ativo(update, ctx, raw, COMPRA_ATIVO_CONFIRM, COMPRA_TIPO, prompt_compra_tipo)


async def compra_ativo_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    return await confirmar_ativo_generico(update, ctx, prompt_compra_tipo)


async def compra_tipo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    ctx.user_data["tipo_investimento"] = raw if raw in TIPOS_INVESTIMENTO else "Outro"
    await update.message.reply_text(
        "🏦 *Corretora:*",
        reply_markup=teclado_lista(get_corretoras_ordenadas(), incluir_outra=True),
        parse_mode="Markdown",
    )
    return COMPRA_CORRETORA


async def compra_corretora(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    if raw == "Outra":
        await update.message.reply_text("Digite o nome da corretora:", reply_markup=teclado_cancelar())
        return COMPRA_CORRETORA_CUSTOM
    ctx.user_data["corretora"] = normalizar_corretora(raw)
    await update.message.reply_text("🔢 *Quantidade comprada:*", reply_markup=teclado_cancelar(), parse_mode="Markdown")
    return COMPRA_QUANTIDADE


async def compra_corretora_custom(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    ctx.user_data["corretora"] = normalizar_corretora(raw)
    await update.message.reply_text("🔢 *Quantidade comprada:*", reply_markup=teclado_cancelar(), parse_mode="Markdown")
    return COMPRA_QUANTIDADE


async def compra_quantidade(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    quantidade = parse_decimal(raw)
    if quantidade is None or quantidade <= 0:
        await update.message.reply_text("Digite uma quantidade maior que zero:")
        return COMPRA_QUANTIDADE
    ctx.user_data["quantidade"] = qtd(quantidade)
    await update.message.reply_text("💵 *Preço unitário:*", reply_markup=teclado_cancelar(), parse_mode="Markdown")
    return COMPRA_PRECO


async def compra_preco(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    preco = parse_decimal(raw)
    if preco is None or preco < 0:
        await update.message.reply_text("Digite um preço válido:")
        return COMPRA_PRECO
    ctx.user_data["preco_unitario"] = preco
    ctx.user_data["valor_total"] = moeda(ctx.user_data["quantidade"] * preco)
    await update.message.reply_text(
        f"✅ *Confirmar compra*\n\n"
        f"Data: {data_br(ctx.user_data['data'])}\n"
        f"Ativo: *{ctx.user_data['ativo']}*\n"
        f"Tipo: {ctx.user_data['tipo_investimento']}\n"
        f"Corretora: {ctx.user_data['corretora']}\n"
        f"Quantidade: {fmt_decimal(ctx.user_data['quantidade'], 8)}\n"
        f"Preço unitário: {fmt_moeda(preco)}\n"
        f"Valor total: *{fmt_moeda(ctx.user_data['valor_total'])}*",
        reply_markup=teclado_confirmar(SALVAR_COMPRA_BTN),
        parse_mode="Markdown",
    )
    return COMPRA_CONFIRMAR


async def compra_confirmar(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    if raw != SALVAR_COMPRA_BTN:
        await update.message.reply_text("Use o botão para salvar ou cancelar.")
        return COMPRA_CONFIRMAR
    novo_id = inserir_compra(ctx.user_data)
    await update.message.reply_text(
        f"Compra #{novo_id} salva: *{ctx.user_data['ativo']}* | "
        f"{fmt_decimal(ctx.user_data['quantidade'], 8)} un. | {fmt_moeda(ctx.user_data['valor_total'])}",
        parse_mode="Markdown",
    )
    return await voltar_menu(update, ctx)


async def venda_inicio(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    await pedir_data(update, "📅 *Data da venda*")
    return VENDA_DATA


async def venda_data(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    data = normalizar_data(raw)
    if not data:
        await update.message.reply_text("Data inválida. Tente novamente:")
        return VENDA_DATA
    ctx.user_data["data"] = data
    await update.message.reply_text("🏷 *Ticker ou nome do ativo vendido:*", reply_markup=teclado_cancelar(), parse_mode="Markdown")
    return VENDA_ATIVO


async def prompt_venda_quantidade(update, ctx):
    disponivel, _, _ = posicao_ativo_ate(ctx.user_data["ativo"], ctx.user_data["data"])
    await update.message.reply_text(
        f"🔢 *Quantidade vendida de {ctx.user_data['ativo']}:*\n"
        f"Disponível na data: *{fmt_decimal(disponivel, 8)}*",
        reply_markup=teclado_cancelar(),
        parse_mode="Markdown",
    )


async def venda_ativo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    return await resolver_ativo(update, ctx, raw, VENDA_ATIVO_CONFIRM, VENDA_QUANTIDADE, prompt_venda_quantidade)


async def venda_ativo_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    return await confirmar_ativo_generico(update, ctx, prompt_venda_quantidade)


async def venda_quantidade(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    quantidade = parse_decimal(raw)
    if quantidade is None or quantidade <= 0:
        await update.message.reply_text("Digite uma quantidade maior que zero:")
        return VENDA_QUANTIDADE
    disponivel, _, _ = posicao_ativo_ate(ctx.user_data["ativo"], ctx.user_data["data"])
    if quantidade > disponivel:
        await update.message.reply_text(
            f"Não dá para registrar essa venda.\n\n"
            f"Ativo: *{ctx.user_data['ativo']}*\n"
            f"Disponível na data: *{fmt_decimal(disponivel, 8)}*\n"
            f"Quantidade informada: *{fmt_decimal(quantidade, 8)}*\n\n"
            f"Digite uma quantidade menor ou cancele.",
            reply_markup=teclado_cancelar(),
            parse_mode="Markdown",
        )
        return VENDA_QUANTIDADE
    ctx.user_data["quantidade"] = qtd(quantidade)
    await update.message.reply_text("💵 *Preço de venda unitário:*", reply_markup=teclado_cancelar(), parse_mode="Markdown")
    return VENDA_PRECO


async def venda_preco(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    preco = parse_decimal(raw)
    if preco is None or preco < 0:
        await update.message.reply_text("Digite um preço válido:")
        return VENDA_PRECO
    preview, disponivel = calcular_venda_preview(
        ctx.user_data["ativo"], ctx.user_data["data"], ctx.user_data["quantidade"], preco
    )
    if preview is None:
        await update.message.reply_text(
            f"Venda bloqueada: posição disponível de {ctx.user_data['ativo']} é "
            f"{fmt_decimal(disponivel, 8)}.",
            reply_markup=teclado_cancelar(),
        )
        return VENDA_QUANTIDADE
    ctx.user_data["preco_unitario"] = preco
    ctx.user_data.update(preview)
    await update.message.reply_text(
        "🏦 *Corretora onde a venda foi executada:*",
        reply_markup=teclado_lista(get_corretoras_ordenadas(), incluir_outra=True),
        parse_mode="Markdown",
    )
    return VENDA_CORRETORA


async def venda_corretora(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    if raw == "Outra":
        await update.message.reply_text("Digite o nome da corretora:", reply_markup=teclado_cancelar())
        return VENDA_CORRETORA_CUSTOM
    ctx.user_data["corretora"] = normalizar_corretora(raw)
    return await venda_confirmacao(update, ctx)


async def venda_corretora_custom(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    ctx.user_data["corretora"] = normalizar_corretora(raw)
    return await venda_confirmacao(update, ctx)


async def venda_confirmacao(update, ctx):
    lucro = ctx.user_data["lucro_prejuizo"]
    sinal = "+" if lucro >= 0 else ""
    await update.message.reply_text(
        f"✅ *Confirmar venda*\n\n"
        f"Data: {data_br(ctx.user_data['data'])}\n"
        f"Ativo: *{ctx.user_data['ativo']}*\n"
        f"Corretora: {ctx.user_data['corretora']}\n"
        f"Quantidade: {fmt_decimal(ctx.user_data['quantidade'], 8)}\n"
        f"Preço unitário: {fmt_moeda(ctx.user_data['preco_unitario'])}\n"
        f"Valor total: *{fmt_moeda(ctx.user_data['valor_total'])}*\n"
        f"Custo médio global: {fmt_moeda(ctx.user_data['custo_medio_unitario'])}\n"
        f"Lucro/prejuízo realizado: *{sinal}{fmt_moeda(lucro)}*",
        reply_markup=teclado_confirmar(SALVAR_VENDA_BTN),
        parse_mode="Markdown",
    )
    return VENDA_CONFIRMAR


async def venda_confirmar(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    if raw != SALVAR_VENDA_BTN:
        await update.message.reply_text("Use o botão para salvar ou cancelar.")
        return VENDA_CONFIRMAR
    preview, disponivel = calcular_venda_preview(
        ctx.user_data["ativo"],
        ctx.user_data["data"],
        ctx.user_data["quantidade"],
        ctx.user_data["preco_unitario"],
    )
    if preview is None:
        await update.message.reply_text(
            f"Venda bloqueada: posição disponível de {ctx.user_data['ativo']} é "
            f"{fmt_decimal(disponivel, 8)}.",
            reply_markup=teclado_menu(),
        )
        return ConversationHandler.END
    ctx.user_data.update(preview)
    novo_id = inserir_venda(ctx.user_data)
    sinal = "+" if ctx.user_data["lucro_prejuizo"] >= 0 else ""
    await update.message.reply_text(
        f"Venda #{novo_id} salva: *{ctx.user_data['ativo']}* | "
        f"{fmt_decimal(ctx.user_data['quantidade'], 8)} un. | "
        f"{sinal}{fmt_moeda(ctx.user_data['lucro_prejuizo'])}",
        parse_mode="Markdown",
    )
    return await voltar_menu(update, ctx)


async def provento_inicio(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    await pedir_data(update, "📅 *Data do provento*")
    return PROVENTO_DATA


async def provento_data(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    data = normalizar_data(raw)
    if not data:
        await update.message.reply_text("Data inválida. Tente novamente:")
        return PROVENTO_DATA
    ctx.user_data["data"] = data
    await update.message.reply_text("🏷 *Ticker ou nome do ativo:*", reply_markup=teclado_cancelar(), parse_mode="Markdown")
    return PROVENTO_ATIVO


async def prompt_provento_corretora(update, ctx):
    await update.message.reply_text(
        "🏦 *Corretora:*",
        reply_markup=teclado_lista(get_corretoras_ordenadas(), incluir_outra=True),
        parse_mode="Markdown",
    )


async def provento_ativo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    return await resolver_ativo(update, ctx, raw, PROVENTO_ATIVO_CONFIRM, PROVENTO_CORRETORA, prompt_provento_corretora)


async def provento_ativo_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    return await confirmar_ativo_generico(update, ctx, prompt_provento_corretora)


async def provento_corretora(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    if raw == "Outra":
        await update.message.reply_text("Digite o nome da corretora:", reply_markup=teclado_cancelar())
        return PROVENTO_CORRETORA_CUSTOM
    ctx.user_data["corretora"] = normalizar_corretora(raw)
    await update.message.reply_text(
        "💰 *Tipo de provento:*",
        reply_markup=teclado_lista(TIPOS_PROVENTO),
        parse_mode="Markdown",
    )
    return PROVENTO_TIPO


async def provento_corretora_custom(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    ctx.user_data["corretora"] = normalizar_corretora(raw)
    await update.message.reply_text(
        "💰 *Tipo de provento:*",
        reply_markup=teclado_lista(TIPOS_PROVENTO),
        parse_mode="Markdown",
    )
    return PROVENTO_TIPO


async def provento_tipo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    ctx.user_data["tipo_provento"] = raw if raw in TIPOS_PROVENTO else "Rendimento"
    await update.message.reply_text("💵 *Valor recebido:*", reply_markup=teclado_cancelar(), parse_mode="Markdown")
    return PROVENTO_VALOR


async def provento_valor(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    valor = parse_decimal(raw)
    if valor is None or valor < 0:
        await update.message.reply_text("Digite um valor válido:")
        return PROVENTO_VALOR
    ctx.user_data["valor_recebido"] = moeda(valor)
    await update.message.reply_text(
        f"✅ *Confirmar provento*\n\n"
        f"Data: {data_br(ctx.user_data['data'])}\n"
        f"Ativo: *{ctx.user_data['ativo']}*\n"
        f"Corretora: {ctx.user_data['corretora']}\n"
        f"Tipo: {ctx.user_data['tipo_provento']}\n"
        f"Valor: *{fmt_moeda(ctx.user_data['valor_recebido'])}*",
        reply_markup=teclado_confirmar(SALVAR_PROVENTO_BTN),
        parse_mode="Markdown",
    )
    return PROVENTO_CONFIRMAR


async def provento_confirmar(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    if raw != SALVAR_PROVENTO_BTN:
        await update.message.reply_text("Use o botão para salvar ou cancelar.")
        return PROVENTO_CONFIRMAR
    novo_id = inserir_provento(ctx.user_data)
    await update.message.reply_text(
        f"Provento #{novo_id} salvo: *{ctx.user_data['ativo']}* | {fmt_moeda(ctx.user_data['valor_recebido'])}",
        parse_mode="Markdown",
    )
    return await voltar_menu(update, ctx)


LIMITE_TELEGRAM = 3500


def paginar_linhas(linhas, limite=LIMITE_TELEGRAM):
    paginas = []
    atual = []
    tamanho = 0
    for linha in linhas:
        tam = len(linha) + 1
        if atual and tamanho + tam > limite:
            paginas.append("\n".join(atual))
            atual = []
            tamanho = 0
        atual.append(linha)
        tamanho += tam
    if atual:
        paginas.append("\n".join(atual))
    return paginas


async def enviar_paginado(update, linhas, reply_markup=None, parse_mode="Markdown"):
    paginas = paginar_linhas(linhas)
    for i, pagina in enumerate(paginas):
        prefixo = f"📄 _Página {i + 1}/{len(paginas)}_\n\n" if len(paginas) > 1 else ""
        await update.message.reply_text(
            prefixo + pagina,
            reply_markup=reply_markup if i == len(paginas) - 1 else None,
            parse_mode=parse_mode,
        )


async def relatorios_inicio(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    pos = calcular_posicoes()
    vendas = listar_vendas()
    proventos = listar_proventos()
    custo = sum((p["custo_total"] for p in pos.values()), Decimal("0"))
    lucro = sum((Decimal(v["lucro_prejuizo"]) for v in vendas), Decimal("0"))
    prov = sum((Decimal(p["valor_recebido"]) for p in proventos), Decimal("0"))
    ativos_com_posicao = sum(1 for p in pos.values() if p["quantidade"] > 0)
    await update.message.reply_text(
        f"📈 *Relatórios*\n\n"
        f"Ativos em carteira: *{ativos_com_posicao}*\n"
        f"Patrimônio a preço de custo: *{fmt_moeda(custo)}*\n"
        f"Lucro/prejuízo realizado: *{fmt_moeda(lucro)}*\n"
        f"Proventos recebidos: *{fmt_moeda(prov)}*",
        reply_markup=teclado_relatorios(),
        parse_mode="Markdown",
    )
    return RELATORIOS_MENU


async def relatorio_posicao(update):
    pos = [p for p in calcular_posicoes().values() if p["quantidade"] > 0]
    pos.sort(key=lambda x: x["custo_total"], reverse=True)
    if not pos:
        await update.message.reply_text("Nenhuma posição em aberto.", reply_markup=teclado_relatorios())
        return
    linhas = ["📌 *Posição atual por ativo*\n"]
    for p in pos:
        linhas.append(
            f"*{p['ativo']}* ({p['tipo_investimento']})\n"
            f"Qtd: {fmt_decimal(p['quantidade'], 8)} | "
            f"Custo médio: {fmt_moeda(p['custo_medio'])} | "
            f"Custo total: *{fmt_moeda(p['custo_total'])}*\n"
            f"Realizado: {fmt_moeda(p['lucro_realizado'])} | Proventos: {fmt_moeda(p['proventos'])}\n"
        )
    await enviar_paginado(update, linhas, reply_markup=teclado_relatorios())


def posicao_por_corretora():
    pos_global = calcular_posicoes()
    linhas = defaultdict(lambda: {"quantidade": Decimal("0"), "custo_total": Decimal("0"), "tipo": ""})
    for c in listar_compras():
        key = (c["corretora"], c["ativo"])
        linhas[key]["quantidade"] += Decimal(c["quantidade"])
        linhas[key]["tipo"] = c["tipo_investimento"]
    for v in listar_vendas():
        key = (v["corretora"], v["ativo"])
        linhas[key]["quantidade"] -= Decimal(v["quantidade"])
    itens = []
    for (corretora, ativo), d in linhas.items():
        if d["quantidade"] == 0:
            continue
        custo_medio = pos_global.get(ativo, {}).get("custo_medio", Decimal("0"))
        d["custo_total"] = moeda(d["quantidade"] * custo_medio)
        d["corretora"] = corretora
        d["ativo"] = ativo
        itens.append(d)
    return sorted(itens, key=lambda x: (x["corretora"], x["ativo"]))


async def relatorio_corretora(update):
    rows = posicao_por_corretora()
    if not rows:
        await update.message.reply_text("Nenhuma posição por corretora.", reply_markup=teclado_relatorios())
        return
    linhas = ["🏦 *Posição por corretora*\n"]
    for r in rows:
        linhas.append(
            f"*{r['corretora']}* | {r['ativo']}\n"
            f"Qtd: {fmt_decimal(r['quantidade'], 8)} | Custo estimado: {fmt_moeda(r['custo_total'])}\n"
        )
    await enviar_paginado(update, linhas, reply_markup=teclado_relatorios())


async def relatorio_tipo(update):
    grupos = defaultdict(lambda: {"ativos": 0, "custo": Decimal("0"), "proventos": Decimal("0"), "realizado": Decimal("0")})
    for p in calcular_posicoes().values():
        if p["quantidade"] <= 0 and p["proventos"] == 0 and p["lucro_realizado"] == 0:
            continue
        tipo = p["tipo_investimento"] or "Sem tipo"
        grupos[tipo]["ativos"] += 1 if p["quantidade"] > 0 else 0
        grupos[tipo]["custo"] += p["custo_total"]
        grupos[tipo]["proventos"] += p["proventos"]
        grupos[tipo]["realizado"] += p["lucro_realizado"]
    linhas = ["🏷 *Resumo por tipo*\n"]
    for tipo, d in sorted(grupos.items(), key=lambda x: x[1]["custo"], reverse=True):
        linhas.append(
            f"*{tipo}*\n"
            f"Ativos: {d['ativos']} | Custo: {fmt_moeda(d['custo'])} | "
            f"Realizado: {fmt_moeda(d['realizado'])} | Proventos: {fmt_moeda(d['proventos'])}\n"
        )
    await enviar_paginado(update, linhas, reply_markup=teclado_relatorios())


async def relatorio_vendas(update):
    vendas = sorted(listar_vendas(), key=lambda x: (data_obj(x["data"]), int(x["id"])), reverse=True)
    if not vendas:
        await update.message.reply_text("Nenhuma venda registrada.", reply_markup=teclado_relatorios())
        return
    linhas = ["📉 *Histórico de vendas*\n"]
    for v in vendas:
        sinal = "+" if Decimal(v["lucro_prejuizo"]) >= 0 else ""
        linhas.append(
            f"#{v['id']} | {data_br(v['data'])} | *{v['ativo']}* | {v['corretora']}\n"
            f"Qtd: {fmt_decimal(v['quantidade'], 8)} | Total: {fmt_moeda(v['valor_total'])} | "
            f"{sinal}{fmt_moeda(v['lucro_prejuizo'])}\n"
        )
    await enviar_paginado(update, linhas, reply_markup=teclado_relatorios())


async def relatorio_proventos(update):
    proventos = sorted(listar_proventos(), key=lambda x: (data_obj(x["data"]), int(x["id"])), reverse=True)
    if not proventos:
        await update.message.reply_text("Nenhum provento registrado.", reply_markup=teclado_relatorios())
        return
    total = sum((Decimal(p["valor_recebido"]) for p in proventos), Decimal("0"))
    linhas = [f"💰 *Proventos recebidos* | Total: *{fmt_moeda(total)}*\n"]
    for p in proventos:
        linhas.append(
            f"#{p['id']} | {data_br(p['data'])} | *{p['ativo']}* | {p['tipo_provento']}\n"
            f"{p['corretora']} | {fmt_moeda(p['valor_recebido'])}\n"
        )
    await enviar_paginado(update, linhas, reply_markup=teclado_relatorios())


def evolucao_patrimonial():
    eventos = []
    for c in listar_compras():
        eventos.append((data_obj(c["data"]), "compra", c))
    for v in listar_vendas():
        eventos.append((data_obj(v["data"]), "venda", v))
    eventos.sort(key=lambda x: (x[0], 0 if x[1] == "compra" else 1, int(x[2]["id"])))
    qtds = defaultdict(Decimal)
    custos = defaultdict(Decimal)
    por_data = {}
    for data, tipo, row in eventos:
        ativo = row["ativo"]
        if tipo == "compra":
            qtds[ativo] += Decimal(row["quantidade"])
            custos[ativo] += Decimal(row["valor_total"])
        else:
            quantidade = Decimal(row["quantidade"])
            custo_medio = Decimal("0") if qtds[ativo] == 0 else custos[ativo] / qtds[ativo]
            custo = moeda(custo_medio * quantidade)
            qtds[ativo] -= quantidade
            custos[ativo] -= custo
            if qtds[ativo] == 0:
                custos[ativo] = Decimal("0")
        por_data[data] = moeda(sum(custos.values(), Decimal("0")))
    return sorted(por_data.items())


async def relatorio_evolucao(update):
    evolucao = evolucao_patrimonial()
    if not evolucao:
        await update.message.reply_text("Ainda não há evolução patrimonial para mostrar.", reply_markup=teclado_relatorios())
        return
    linhas = ["📅 *Evolução patrimonial a preço de custo*\n"]
    for data, valor in evolucao[-30:]:
        linhas.append(f"{data.strftime('%d/%m/%Y')}: *{fmt_moeda(valor)}*")
    await enviar_paginado(update, linhas, reply_markup=teclado_relatorios())


async def relatorios_resposta(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    txt = update.message.text.strip()
    if txt == "🔙 Voltar":
        return await voltar_menu(update, ctx)
    if txt == "📌 Posição Atual":
        await relatorio_posicao(update)
    elif txt == "🏦 Por Corretora":
        await relatorio_corretora(update)
    elif txt == "🏷 Por Tipo":
        await relatorio_tipo(update)
    elif txt == "📉 Vendas":
        await relatorio_vendas(update)
    elif txt == "💰 Proventos":
        await relatorio_proventos(update)
    elif txt == "📅 Evolução":
        await relatorio_evolucao(update)
    else:
        await update.message.reply_text("Use os botões abaixo.", reply_markup=teclado_relatorios())
    return RELATORIOS_MENU


TIPOS_EDITAR = {"Compras": "compras", "Vendas": "vendas", "Proventos": "proventos"}
REGISTROS_POR_PAGINA = 15


def listar_registros(tipo):
    tabelas = {"compras": "compras", "vendas": "vendas", "proventos": "proventos"}
    return fetch_all(f"SELECT * FROM {tabelas[tipo]} ORDER BY data DESC, id DESC")


def buscar_registro(tipo, registro_id):
    return fetch_one(f"SELECT * FROM {tipo} WHERE id=%s", (registro_id,))


def editar_campos(tipo):
    if tipo == "compras":
        return {
            "data": "📅 Data",
            "ativo": "🏷 Ativo",
            "tipo_investimento": "📚 Tipo",
            "corretora": "🏦 Corretora",
            "quantidade": "🔢 Quantidade",
            "preco_unitario": "💵 Preço unitário",
            "observacao": "📝 Observação",
            "deletar": "🗑 Deletar",
        }
    if tipo == "vendas":
        return {
            "data": "📅 Data",
            "ativo": "🏷 Ativo",
            "corretora": "🏦 Corretora",
            "quantidade": "🔢 Quantidade",
            "preco_unitario": "💵 Preço unitário",
            "deletar": "🗑 Deletar",
        }
    return {
        "data": "📅 Data",
        "ativo": "🏷 Ativo",
        "corretora": "🏦 Corretora",
        "tipo_provento": "💰 Tipo",
        "valor_recebido": "💵 Valor",
        "deletar": "🗑 Deletar",
    }


def resumo_registro(tipo, r):
    if tipo == "compras":
        return f"#{r['id']} | {data_br(r['data'])} | {r['ativo']} | {fmt_decimal(r['quantidade'], 8)} | {fmt_moeda(r['valor_total'])}"
    if tipo == "vendas":
        return f"#{r['id']} | {data_br(r['data'])} | {r['ativo']} | {fmt_decimal(r['quantidade'], 8)} | {fmt_moeda(r['lucro_prejuizo'])}"
    return f"#{r['id']} | {data_br(r['data'])} | {r['ativo']} | {r['tipo_provento']} | {fmt_moeda(r['valor_recebido'])}"


async def editar_inicio(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ctx.user_data.clear()
    await update.message.reply_text(
        "✏️ *Editar registro*\nEscolha uma categoria:",
        reply_markup=ReplyKeyboardMarkup([["Compras"], ["Vendas"], ["Proventos"], [CANCELAR_BTN]], resize_keyboard=True),
        parse_mode="Markdown",
    )
    return EDITAR_TIPO


def montar_pagina_edicao(tipo, pagina):
    regs = listar_registros(tipo)
    total_paginas = max(1, (len(regs) + REGISTROS_POR_PAGINA - 1) // REGISTROS_POR_PAGINA)
    pagina = max(0, min(pagina, total_paginas - 1))
    fatia = regs[pagina * REGISTROS_POR_PAGINA : (pagina + 1) * REGISTROS_POR_PAGINA]
    titulo = {"compras": "Compras", "vendas": "Vendas", "proventos": "Proventos"}[tipo]
    linhas = [f"✏️ *{titulo}* — Página {pagina + 1}/{total_paginas}\n"]
    linhas += [resumo_registro(tipo, r) for r in fatia]
    linhas.append("\nDigite o ID do registro:")
    nav = []
    if pagina > 0:
        nav.append("⬅️ Anterior")
    if pagina < total_paginas - 1:
        nav.append("➡️ Próxima")
    teclado = []
    if nav:
        teclado.append(nav)
    teclado.append([CANCELAR_BTN])
    return "\n".join(linhas), teclado, pagina


async def editar_tipo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    tipo = TIPOS_EDITAR.get(raw)
    if not tipo:
        await update.message.reply_text("Escolha uma categoria do teclado:")
        return EDITAR_TIPO
    ctx.user_data["editar_tipo"] = tipo
    ctx.user_data["editar_pagina"] = 0
    texto, teclado, _ = montar_pagina_edicao(tipo, 0)
    await update.message.reply_text(
        texto,
        reply_markup=ReplyKeyboardMarkup(teclado, resize_keyboard=True),
        parse_mode="Markdown",
    )
    return EDITAR_PAGINA


async def editar_pagina(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    tipo = ctx.user_data["editar_tipo"]
    pagina = ctx.user_data.get("editar_pagina", 0)
    if raw == "➡️ Próxima":
        pagina += 1
    elif raw == "⬅️ Anterior":
        pagina -= 1
    else:
        try:
            registro_id = int(raw)
        except ValueError:
            await update.message.reply_text("Digite o ID do registro ou use os botões.")
            return EDITAR_PAGINA
        registro = buscar_registro(tipo, registro_id)
        if not registro:
            await update.message.reply_text("ID não encontrado. Tente novamente:")
            return EDITAR_PAGINA
        ctx.user_data["editar_id"] = registro_id
        labels = editar_campos(tipo)
        teclado = [[v] for v in labels.values()]
        teclado.append([CANCELAR_BTN])
        await update.message.reply_text(
            f"{resumo_registro(tipo, registro)}\n\nQual campo deseja editar?",
            reply_markup=ReplyKeyboardMarkup(teclado, resize_keyboard=True, one_time_keyboard=True),
        )
        return EDITAR_CAMPO
    ctx.user_data["editar_pagina"] = pagina
    texto, teclado, pagina = montar_pagina_edicao(tipo, pagina)
    ctx.user_data["editar_pagina"] = pagina
    await update.message.reply_text(texto, reply_markup=ReplyKeyboardMarkup(teclado, resize_keyboard=True), parse_mode="Markdown")
    return EDITAR_PAGINA


async def editar_campo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    tipo = ctx.user_data["editar_tipo"]
    labels = editar_campos(tipo)
    campo = next((k for k, v in labels.items() if v == raw), None)
    if not campo:
        await update.message.reply_text("Escolha um campo do teclado:")
        return EDITAR_CAMPO
    ctx.user_data["editar_campo"] = campo
    if campo == "deletar":
        await update.message.reply_text(
            "Esta ação remove o registro. Confirma?",
            reply_markup=ReplyKeyboardMarkup([[SIM_DELETAR_BTN], [NAO_DELETAR_BTN]], resize_keyboard=True),
        )
        return EDITAR_VALOR
    if campo == "tipo_investimento":
        await update.message.reply_text("Novo tipo:", reply_markup=teclado_lista(TIPOS_INVESTIMENTO))
        return EDITAR_SELECT
    if campo == "tipo_provento":
        await update.message.reply_text("Novo tipo de provento:", reply_markup=teclado_lista(TIPOS_PROVENTO))
        return EDITAR_SELECT
    if campo == "corretora":
        await update.message.reply_text("Nova corretora:", reply_markup=teclado_lista(get_corretoras_ordenadas(), incluir_outra=True))
        return EDITAR_SELECT
    prompts = {
        "data": "Nova data ou 0 para hoje:",
        "ativo": "Novo ticker ou nome do ativo:",
        "quantidade": "Nova quantidade:",
        "preco_unitario": "Novo preço unitário:",
        "valor_recebido": "Novo valor recebido:",
        "observacao": "Nova observação:",
    }
    await update.message.reply_text(prompts[campo], reply_markup=teclado_cancelar())
    return EDITAR_VALOR


def atualizar_valor_total(tipo, registro_id):
    if tipo not in ("compras", "vendas"):
        return
    r = buscar_registro(tipo, registro_id)
    valor_total = moeda(Decimal(r["quantidade"]) * Decimal(r["preco_unitario"]))
    with conectar() as conn:
        with conn.cursor() as cur:
            cur.execute(f"UPDATE {tipo} SET valor_total=%s WHERE id=%s", (valor_total, registro_id))
        conn.commit()


def aplicar_update_seguro(tipo, registro_id, campo, valor):
    antigo = buscar_registro(tipo, registro_id)
    ativos_afetados = {antigo.get("ativo")}
    try:
        with conectar() as conn:
            with conn.cursor() as cur:
                cur.execute(f"UPDATE {tipo} SET {campo}=%s WHERE id=%s", (valor, registro_id))
            conn.commit()
        if campo in ("quantidade", "preco_unitario"):
            atualizar_valor_total(tipo, registro_id)
        novo = buscar_registro(tipo, registro_id)
        ativos_afetados.add(novo.get("ativo"))
        recalcular_vendas([a for a in ativos_afetados if a])
    except Exception:
        with conectar() as conn:
            with conn.cursor() as cur:
                cols = [k for k in antigo.keys() if k not in ("id", "criado_em")]
                sets = ", ".join(f"{c}=%s" for c in cols)
                cur.execute(f"UPDATE {tipo} SET {sets} WHERE id=%s", [antigo[c] for c in cols] + [registro_id])
            conn.commit()
        recalcular_vendas([a for a in ativos_afetados if a])
        raise


def deletar_seguro(tipo, registro_id):
    antigo = buscar_registro(tipo, registro_id)
    if not antigo:
        return
    cols = [k for k in antigo.keys() if k not in ("id", "criado_em")]
    try:
        with conectar() as conn:
            with conn.cursor() as cur:
                cur.execute(f"DELETE FROM {tipo} WHERE id=%s", (registro_id,))
            conn.commit()
        recalcular_vendas([antigo.get("ativo")])
    except Exception:
        with conectar() as conn:
            with conn.cursor() as cur:
                col_list = ", ".join(["id"] + cols)
                placeholders = ", ".join(["%s"] * (len(cols) + 1))
                cur.execute(
                    f"INSERT INTO {tipo} ({col_list}) VALUES ({placeholders})",
                    [registro_id] + [antigo[c] for c in cols],
                )
            conn.commit()
        recalcular_vendas([antigo.get("ativo")])
        raise


async def editar_select(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN:
        return await cancelar(update, ctx)
    tipo = ctx.user_data["editar_tipo"]
    campo = ctx.user_data["editar_campo"]
    if campo == "corretora":
        if raw == "Outra":
            await update.message.reply_text("Digite o nome da corretora:", reply_markup=teclado_cancelar())
            return EDITAR_VALOR
        valor = normalizar_corretora(raw)
    else:
        valor = raw
    try:
        aplicar_update_seguro(tipo, ctx.user_data["editar_id"], campo, valor)
    except Exception as exc:
        await update.message.reply_text(f"Não foi possível aplicar a edição: {exc}", reply_markup=teclado_menu())
        return ConversationHandler.END
    await update.message.reply_text("Registro atualizado.")
    return await voltar_menu(update, ctx)


async def editar_valor(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == CANCELAR_BTN or raw == NAO_DELETAR_BTN:
        return await cancelar(update, ctx)
    tipo = ctx.user_data["editar_tipo"]
    campo = ctx.user_data["editar_campo"]
    registro_id = ctx.user_data["editar_id"]
    if campo == "deletar":
        if raw != SIM_DELETAR_BTN:
            await update.message.reply_text("Remoção cancelada.")
            return await voltar_menu(update, ctx)
        try:
            deletar_seguro(tipo, registro_id)
        except Exception as exc:
            await update.message.reply_text(f"Não foi possível deletar: {exc}", reply_markup=teclado_menu())
            return ConversationHandler.END
        await update.message.reply_text("Registro deletado.")
        return await voltar_menu(update, ctx)
    if campo == "data":
        valor = normalizar_data(raw)
        if not valor:
            await update.message.reply_text("Data inválida:")
            return EDITAR_VALOR
    elif campo == "ativo":
        valor = normalizar_ativo(raw)
    elif campo in ("quantidade", "preco_unitario", "valor_recebido"):
        valor = parse_decimal(raw)
        if valor is None or valor < 0 or (campo == "quantidade" and valor <= 0):
            await update.message.reply_text("Digite um número válido:")
            return EDITAR_VALOR
        valor = qtd(valor) if campo == "quantidade" else valor
    elif campo == "corretora":
        valor = normalizar_corretora(raw)
    else:
        valor = raw
    try:
        aplicar_update_seguro(tipo, registro_id, campo, valor)
    except Exception as exc:
        await update.message.reply_text(f"Não foi possível aplicar a edição: {exc}", reply_markup=teclado_menu())
        return ConversationHandler.END
    await update.message.reply_text("Registro atualizado.")
    return await voltar_menu(update, ctx)


async def gerar_menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📊 *Gerar arquivos*\n\nEscolha uma opção:",
        reply_markup=teclado_gerar(),
        parse_mode="Markdown",
    )
    return GERAR_MENU


def adicionar_tabela(ws, headers, rows, start_row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = cell.font.copy(bold=True)
    for r_idx, row in enumerate(rows, start_row + 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = 18


async def gerar_dashboard(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Gerando dashboard, aguarde...")
    import openpyxl
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    try:
        recalcular_vendas()
    except ValueError as exc:
        await update.message.reply_text(
            f'Não foi possível gerar o dashboard porque há inconsistência histórica: {exc}',
            reply_markup=teclado_gerar(),
        )
        return
    pos = calcular_posicoes()
    compras = listar_compras()
    vendas = listar_vendas()
    proventos = listar_proventos()
    evolucao = evolucao_patrimonial()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Dashboard de Investimentos"
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells("A1:F1")
    indicadores = [
        ("Ativos em carteira", sum(1 for p in pos.values() if p["quantidade"] > 0)),
        ("Patrimônio a preço de custo", float(sum((p["custo_total"] for p in pos.values()), Decimal("0")))),
        ("Lucro/prejuízo realizado", float(sum((Decimal(v["lucro_prejuizo"]) for v in vendas), Decimal("0")))),
        ("Proventos recebidos", float(sum((Decimal(p["valor_recebido"]) for p in proventos), Decimal("0")))),
        ("Compras registradas", len(compras)),
        ("Vendas registradas", len(vendas)),
    ]
    for i, (label, valor) in enumerate(indicadores, 3):
        ws.cell(i, 1, label).font = Font(bold=True)
        ws.cell(i, 2, valor)
        if isinstance(valor, float):
            ws.cell(i, 2).number_format = '"R$" #,##0.00'

    pos_rows = []
    for p in sorted(pos.values(), key=lambda x: x["custo_total"], reverse=True):
        if p["quantidade"] <= 0:
            continue
        pos_rows.append(
            [
                p["ativo"],
                p["tipo_investimento"],
                float(p["quantidade"]),
                float(p["custo_medio"]),
                float(p["custo_total"]),
                float(p["lucro_realizado"]),
                float(p["proventos"]),
            ]
        )
    wpos = wb.create_sheet("Posição Atual")
    adicionar_tabela(
        wpos,
        ["Ativo", "Tipo", "Quantidade", "Custo médio", "Custo total", "Realizado", "Proventos"],
        pos_rows,
    )
    for row in wpos.iter_rows(min_row=2, min_col=4, max_col=7):
        for cell in row:
            cell.number_format = '"R$" #,##0.00'

    wcor = wb.create_sheet("Por Corretora")
    cor_rows = [
        [r["corretora"], r["ativo"], float(r["quantidade"]), float(r["custo_total"])]
        for r in posicao_por_corretora()
    ]
    adicionar_tabela(wcor, ["Corretora", "Ativo", "Quantidade", "Custo estimado"], cor_rows)
    for cell in wcor["D"][1:]:
        cell.number_format = '"R$" #,##0.00'

    wtipo = wb.create_sheet("Por Tipo")
    grupos = defaultdict(lambda: {"custo": Decimal("0"), "realizado": Decimal("0"), "proventos": Decimal("0")})
    for p in pos.values():
        tipo = p["tipo_investimento"] or "Sem tipo"
        grupos[tipo]["custo"] += p["custo_total"]
        grupos[tipo]["realizado"] += p["lucro_realizado"]
        grupos[tipo]["proventos"] += p["proventos"]
    tipo_rows = [[k, float(v["custo"]), float(v["realizado"]), float(v["proventos"])] for k, v in grupos.items()]
    adicionar_tabela(wtipo, ["Tipo", "Custo atual", "Realizado", "Proventos"], tipo_rows)
    if tipo_rows:
        chart = BarChart()
        chart.title = "Custo atual por tipo"
        data = Reference(wtipo, min_col=2, min_row=1, max_row=len(tipo_rows) + 1)
        cats = Reference(wtipo, min_col=1, min_row=2, max_row=len(tipo_rows) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        wtipo.add_chart(chart, "F2")

    wv = wb.create_sheet("Vendas")
    venda_rows = [
        [
            data_br(v["data"]),
            v["ativo"],
            v.get("tipo_investimento") or "",
            v["corretora"],
            float(v["quantidade"]),
            float(v["preco_unitario"]),
            float(v["valor_total"]),
            float(v["custo_medio_unitario"]),
            float(v["lucro_prejuizo"]),
        ]
        for v in vendas
    ]
    adicionar_tabela(
        wv,
        ["Data", "Ativo", "Tipo", "Corretora", "Quantidade", "Preço venda", "Valor venda", "Custo médio", "Lucro/prejuízo"],
        venda_rows,
    )

    wp = wb.create_sheet("Proventos")
    prov_rows = [
        [data_br(p["data"]), p["ativo"], p["corretora"], p["tipo_provento"], float(p["valor_recebido"])]
        for p in proventos
    ]
    adicionar_tabela(wp, ["Data", "Ativo", "Corretora", "Tipo", "Valor recebido"], prov_rows)

    wc = wb.create_sheet("Compras")
    compra_rows = [
        [
            data_br(c["data"]),
            c["ativo"],
            c["tipo_investimento"],
            c["corretora"],
            float(c["quantidade"]),
            float(c["preco_unitario"]),
            float(c["valor_total"]),
            c.get("observacao") or "",
        ]
        for c in compras
    ]
    adicionar_tabela(wc, ["Data", "Ativo", "Tipo", "Corretora", "Quantidade", "Preço unitário", "Valor total", "Observação"], compra_rows)

    we = wb.create_sheet("Evolução")
    adicionar_tabela(we, ["Data", "Patrimônio a custo"], [[d.strftime("%d/%m/%Y"), float(v)] for d, v in evolucao])
    if len(evolucao) >= 2:
        chart = LineChart()
        chart.title = "Evolução patrimonial a preço de custo"
        chart.y_axis.title = "R$"
        data = Reference(we, min_col=2, min_row=1, max_row=len(evolucao) + 1)
        cats = Reference(we, min_col=1, min_row=2, max_row=len(evolucao) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        we.add_chart(chart, "D2")

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")
        for col in range(1, min(sheet.max_column, 12) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = max(sheet.column_dimensions[get_column_letter(col)].width or 12, 15)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome = f"dashboard_investimentos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    await update.message.reply_document(document=InputFile(buf, filename=nome), caption="Dashboard gerado.")


def csv_bytes(rows, headers):
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        item = {}
        for h in headers:
            v = row.get(h)
            if isinstance(v, Decimal):
                item[h] = str(v)
            elif hasattr(v, "strftime"):
                item[h] = v.strftime("%Y-%m-%d")
            else:
                item[h] = v
        writer.writerow(item)
    return buf.getvalue().encode("utf-8-sig")


async def exportar_csv(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Gerando exportação CSV...")
    compras = listar_compras()
    vendas = listar_vendas()
    proventos = listar_proventos()
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(
            "compras.csv",
            csv_bytes(
                compras,
                ["id", "data", "ativo", "tipo_investimento", "corretora", "quantidade", "preco_unitario", "valor_total", "observacao"],
            ),
        )
        zf.writestr(
            "vendas.csv",
            csv_bytes(
                vendas,
                [
                    "id",
                    "data",
                    "ativo",
                    "tipo_investimento",
                    "corretora",
                    "quantidade",
                    "preco_unitario",
                    "valor_total",
                    "custo_medio_unitario",
                    "custo_total",
                    "lucro_prejuizo",
                ],
            ),
        )
        zf.writestr(
            "proventos.csv",
            csv_bytes(proventos, ["id", "data", "ativo", "corretora", "tipo_provento", "valor_recebido"]),
        )
    zip_buf.seek(0)
    await update.message.reply_document(
        document=InputFile(zip_buf, filename="investimentos_csv.zip"),
        caption="Exportação gerada.",
    )


async def gerar_script_sql(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    conteudo = (
        "-- Schema do bot de investimentos.\n"
        "-- Cria as tabelas novas do zero e não reaproveita estruturas antigas.\n\n"
        f"{SCHEMA_SQL}\n"
    ).encode("utf-8")
    await update.message.reply_document(
        document=InputFile(io.BytesIO(conteudo), filename="schema_investimentos.sql"),
        caption="Script SQL gerado.",
    )


async def gerar_resposta(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    txt = update.message.text.strip()
    if txt == "🔙 Voltar":
        return await voltar_menu(update, ctx)
    if txt == "📊 Dashboard Excel":
        await gerar_dashboard(update, ctx)
    elif txt == "📂 Exportar CSV":
        await exportar_csv(update, ctx)
    elif txt == "🧱 Script SQL":
        await gerar_script_sql(update, ctx)
    else:
        await update.message.reply_text("Use os botões abaixo.", reply_markup=teclado_gerar())
    return GERAR_MENU


async def menu_botao(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    txt = update.message.text.strip()
    if txt == "📈 Relatórios":
        return await relatorios_inicio(update, ctx)
    if txt == "📊 Gerar Arquivos":
        return await gerar_menu(update, ctx)
    await update.message.reply_text("Escolha uma opção do menu.", reply_markup=teclado_menu())


class PingHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write(b"ok")

    def do_HEAD(self):
        self.send_response(200)
        self.end_headers()

    def log_message(self, *args):
        pass


def iniciar_servidor():
    porta = int(os.environ.get("PORT", "10000"))
    server = HTTPServer(("0.0.0.0", porta), PingHandler)
    threading.Thread(target=server.serve_forever, daemon=True).start()


def exigir_variaveis_ambiente():
    faltando = []
    if not TOKEN:
        faltando.append("TELEGRAM_BOT_TOKEN")
    if not DATABASE_URL:
        faltando.append("DATABASE_URL")
    if faltando:
        raise RuntimeError("Defina as variáveis de ambiente: " + ", ".join(faltando))


def main():
    exigir_variaveis_ambiente()
    inicializar_db()
    iniciar_servidor()
    app = Application.builder().token(TOKEN).build()

    botoes_menu = [
        "🟢 Registrar Compra",
        "🔴 Registrar Venda",
        "💰 Registrar Provento",
        "📈 Relatórios",
        "✏️ Editar Registro",
        "📊 Gerar Arquivos",
    ]
    fallbacks = [
        CommandHandler("cancelar", cancelar),
        MessageHandler(filters.Regex(f"^{re.escape(CANCELAR_BTN)}$"), cancelar),
        MessageHandler(filters.Regex("^(" + "|".join(re.escape(b) for b in botoes_menu) + ")$"), cancelar),
    ]

    conv_compra = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^🟢 Registrar Compra$"), compra_inicio)],
        states={
            COMPRA_DATA: [MessageHandler(filters.TEXT & ~filters.COMMAND, compra_data)],
            COMPRA_ATIVO: [MessageHandler(filters.TEXT & ~filters.COMMAND, compra_ativo)],
            COMPRA_ATIVO_CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, compra_ativo_confirm)],
            COMPRA_TIPO: [MessageHandler(filters.TEXT & ~filters.COMMAND, compra_tipo)],
            COMPRA_CORRETORA: [MessageHandler(filters.TEXT & ~filters.COMMAND, compra_corretora)],
            COMPRA_CORRETORA_CUSTOM: [MessageHandler(filters.TEXT & ~filters.COMMAND, compra_corretora_custom)],
            COMPRA_QUANTIDADE: [MessageHandler(filters.TEXT & ~filters.COMMAND, compra_quantidade)],
            COMPRA_PRECO: [MessageHandler(filters.TEXT & ~filters.COMMAND, compra_preco)],
            COMPRA_CONFIRMAR: [MessageHandler(filters.TEXT & ~filters.COMMAND, compra_confirmar)],
        },
        fallbacks=fallbacks,
    )

    conv_venda = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^🔴 Registrar Venda$"), venda_inicio)],
        states={
            VENDA_DATA: [MessageHandler(filters.TEXT & ~filters.COMMAND, venda_data)],
            VENDA_ATIVO: [MessageHandler(filters.TEXT & ~filters.COMMAND, venda_ativo)],
            VENDA_ATIVO_CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, venda_ativo_confirm)],
            VENDA_QUANTIDADE: [MessageHandler(filters.TEXT & ~filters.COMMAND, venda_quantidade)],
            VENDA_PRECO: [MessageHandler(filters.TEXT & ~filters.COMMAND, venda_preco)],
            VENDA_CORRETORA: [MessageHandler(filters.TEXT & ~filters.COMMAND, venda_corretora)],
            VENDA_CORRETORA_CUSTOM: [MessageHandler(filters.TEXT & ~filters.COMMAND, venda_corretora_custom)],
            VENDA_CONFIRMAR: [MessageHandler(filters.TEXT & ~filters.COMMAND, venda_confirmar)],
        },
        fallbacks=fallbacks,
    )

    conv_provento = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^💰 Registrar Provento$"), provento_inicio)],
        states={
            PROVENTO_DATA: [MessageHandler(filters.TEXT & ~filters.COMMAND, provento_data)],
            PROVENTO_ATIVO: [MessageHandler(filters.TEXT & ~filters.COMMAND, provento_ativo)],
            PROVENTO_ATIVO_CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, provento_ativo_confirm)],
            PROVENTO_CORRETORA: [MessageHandler(filters.TEXT & ~filters.COMMAND, provento_corretora)],
            PROVENTO_CORRETORA_CUSTOM: [MessageHandler(filters.TEXT & ~filters.COMMAND, provento_corretora_custom)],
            PROVENTO_TIPO: [MessageHandler(filters.TEXT & ~filters.COMMAND, provento_tipo)],
            PROVENTO_VALOR: [MessageHandler(filters.TEXT & ~filters.COMMAND, provento_valor)],
            PROVENTO_CONFIRMAR: [MessageHandler(filters.TEXT & ~filters.COMMAND, provento_confirmar)],
        },
        fallbacks=fallbacks,
    )

    conv_editar = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^✏️ Editar Registro$"), editar_inicio)],
        states={
            EDITAR_TIPO: [MessageHandler(filters.TEXT & ~filters.COMMAND, editar_tipo)],
            EDITAR_PAGINA: [MessageHandler(filters.TEXT & ~filters.COMMAND, editar_pagina)],
            EDITAR_CAMPO: [MessageHandler(filters.TEXT & ~filters.COMMAND, editar_campo)],
            EDITAR_VALOR: [MessageHandler(filters.TEXT & ~filters.COMMAND, editar_valor)],
            EDITAR_SELECT: [MessageHandler(filters.TEXT & ~filters.COMMAND, editar_select)],
        },
        fallbacks=fallbacks,
    )

    conv_relatorios = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^📈 Relatórios$"), relatorios_inicio)],
        states={RELATORIOS_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, relatorios_resposta)]},
        fallbacks=fallbacks,
    )

    conv_gerar = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^📊 Gerar Arquivos$"), gerar_menu)],
        states={GERAR_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, gerar_resposta)]},
        fallbacks=fallbacks,
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("exportar", exportar_csv))
    app.add_handler(CommandHandler("schema", gerar_script_sql))
    app.add_handler(conv_compra)
    app.add_handler(conv_venda)
    app.add_handler(conv_provento)
    app.add_handler(conv_editar)
    app.add_handler(conv_relatorios)
    app.add_handler(conv_gerar)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, menu_botao))
    app.run_polling()


if __name__ == "__main__":
    main()



